# eventos/views.py
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse, JsonResponse
from django.urls import reverse
from django.utils import timezone
from django.contrib import messages
from django.db.models import Q, Count
from django.db.models.functions import TruncMonth
from django.template.loader import render_to_string
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from datetime import timedelta, datetime
import json
import pytz
from .utils import get_mexico_timezone, convert_to_mexico_time, format_event_date

# Importaciones para reportes
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

from .models import Evento, Municipio
from .forms import EventoForm, FiltroEventosForm

#importacion Chatbot
from django.conf import settings
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_http_methods
from .chatbot import ChatbotAgenda

# Vista principal - Dashboard
@login_required
def dashboard(request):
    """Dashboard principal con eventos del día y próximos"""
    # Obtener fecha actual en zona horaria de México
    mexico_tz = pytz.timezone('America/Mexico_City')
    ahora_mexico = timezone.now().astimezone(mexico_tz)
    hoy = ahora_mexico.date()
    
    print(f"DEBUG - Fecha/hora actual en México: {ahora_mexico}")
    print(f"DEBUG - Fecha de hoy: {hoy}")
    
    # Obtener todos los eventos de hoy
    eventos_hoy_todos = Evento.objects.filter(
        fecha_evento__date=hoy
    ).select_related('municipio', 'creado_por').order_by('fecha_evento')
    
    print(f"DEBUG - Eventos encontrados para hoy: {eventos_hoy_todos.count()}")
    
    # Actualizar estados automáticamente antes de mostrar
    for evento in eventos_hoy_todos:
        estado_anterior = evento.estado
        evento.actualizar_estado_automatico()
        if estado_anterior != evento.estado:
            print(f"DEBUG - Estado actualizado: {evento.nombre} {estado_anterior} -> {evento.estado}")
    
    # Refrescar desde la base de datos después de actualizar
    eventos_hoy_todos = Evento.objects.filter(
        fecha_evento__date=hoy
    ).select_related('municipio', 'creado_por').order_by('fecha_evento')
    
    # Separar por estados calculados
    eventos_en_curso = eventos_hoy_todos.filter(estado='en_curso')
    eventos_hoy_proximos = eventos_hoy_todos.filter(estado='programado')
    eventos_hoy_finalizados = eventos_hoy_todos.filter(estado='finalizado')
    
    # Eventos próximos (siguientes 7 días, excluyendo hoy)
    fecha_limite = hoy + timedelta(days=7)
    eventos_proximos = Evento.objects.filter(
        fecha_evento__date__gt=hoy,
        fecha_evento__date__lte=fecha_limite
    ).select_related('municipio', 'creado_por').order_by('fecha_evento')[:10]
    
    # También actualizar estados de eventos próximos
    for evento in eventos_proximos:
        evento.actualizar_estado_automatico()
    
    # Debug: Imprimir información para verificar
    print(f"DEBUG - Total eventos hoy: {eventos_hoy_todos.count()}")
    print(f"DEBUG - En curso: {eventos_en_curso.count()}")
    print(f"DEBUG - Próximos hoy: {eventos_hoy_proximos.count()}")
    print(f"DEBUG - Finalizados hoy: {eventos_hoy_finalizados.count()}")
    print(f"DEBUG - Próximos 7 días: {eventos_proximos.count()}")
    
    context = {
        'eventos_hoy_todos': eventos_hoy_todos,
        'eventos_en_curso': eventos_en_curso,
        'eventos_hoy_proximos': eventos_hoy_proximos,
        'eventos_hoy_finalizados': eventos_hoy_finalizados,
        'eventos_proximos': eventos_proximos,
        'fecha_actual': hoy,
        'ahora_mexico': ahora_mexico,
    }
    
    return render(request, 'eventos/dashboard.html', context)

# Vista para crear eventos
@login_required
def crear_evento(request):
    """Formulario para crear nuevos eventos"""
    if request.method == 'POST':
        form = EventoForm(request.POST)
        
        if form.is_valid():
            try:
                # Crear evento sin guardarlo aún
                evento = form.save(commit=False)
                evento.creado_por = request.user
                
                # Guardar el evento
                evento.save()
                
                # Mensaje de éxito personalizado
                messages.success(
                    request, 
                    f'✅ Evento "{evento.nombre}" creado exitosamente para el '
                    f'{evento.get_fecha_mexico().strftime("%d/%m/%Y a las %H:%M")} '
                    f'en {evento.municipio.nombre}.'
                )
                
                return redirect('dashboard')
                    
            except Exception as e:
                # Log del error para debugging
                print(f"Error al crear evento: {e}")
                messages.error(
                    request, 
                    '❌ Ocurrió un error al crear el evento. Por favor, inténtelo nuevamente.'
                )
        else:
            # Mostrar errores específicos del formulario
            error_messages = []
            for field, errors in form.errors.items():
                field_label = form.fields[field].label if field in form.fields else field
                for error in errors:
                    error_messages.append(f"{field_label}: {error}")
            
            if error_messages:
                messages.error(
                    request, 
                    '❌ Por favor corrija los siguientes errores: ' + 
                    ', '.join(error_messages)
                )
            else:
                messages.error(
                    request, 
                    '❌ Por favor corrija los errores en el formulario.'
                )
    else:
        form = EventoForm()
        
        # Pre-llenar algunos campos si vienen por GET
        if request.GET.get('municipio'):
            try:
                municipio_id = int(request.GET.get('municipio'))
                form.initial['municipio'] = municipio_id
            except (ValueError, TypeError):
                pass
        
        if request.GET.get('fecha'):
            try:
                fecha_str = request.GET.get('fecha')
                fecha = datetime.strptime(fecha_str, '%Y-%m-%d')
                form.initial['fecha_evento'] = fecha.replace(hour=9)
            except (ValueError, TypeError):
                pass
    
    context = {
        'form': form,
        'municipios_activos': Municipio.objects.filter(activo=True).count(),
        'hora_actual': timezone.now().astimezone(pytz.timezone('America/Mexico_City')),
    }
    
    return render(request, 'eventos/crear_evento.html', context)

# Vista para finalizar evento manualmente
@login_required
def finalizar_evento(request, pk):
    """Finaliza un evento manualmente"""
    evento = get_object_or_404(Evento, pk=pk)
    
    if request.method == 'POST':
        if evento.puede_finalizar_manualmente:
            evento.finalizar_manualmente()
            messages.success(request, f'Evento "{evento.nombre}" marcado como finalizado.')
        else:
            messages.error(request, 'Este evento no puede ser finalizado manualmente en este momento.')
    
    return redirect('dashboard')

# Vista para cambiar estado de evento
@login_required
def cambiar_estado_evento(request, pk):
    """Cambia el estado de un evento manualmente"""
    evento = get_object_or_404(Evento, pk=pk)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado in ['programado', 'en_curso', 'finalizado', 'cancelado']:
            evento.estado = nuevo_estado
            if nuevo_estado == 'finalizado':
                evento.fecha_finalizacion_manual = timezone.now()
            evento.save()
            messages.success(request, f'Estado del evento cambiado a "{evento.get_estado_display()}".')
        else:
            messages.error(request, 'Estado no válido.')
    
    return redirect('dashboard')

# Vista para forzar actualización de estados
@login_required
def actualizar_estados_eventos(request):
    """Actualiza todos los estados de eventos automáticamente"""
    if request.method == 'POST':
        eventos = Evento.objects.filter(fecha_finalizacion_manual__isnull=True)
        contador = 0
        
        for evento in eventos:
            estado_anterior = evento.estado
            evento.actualizar_estado_automatico()
            if evento.estado != estado_anterior:
                contador += 1
        
        messages.success(request, f'Se actualizaron {contador} eventos automáticamente.')
    
    return redirect('dashboard')

# Vista para editar eventos
# eventos/views.py - Vista corregida para editar_evento

# Reemplaza la función editar_evento existente en eventos/views.py con esta versión:

@login_required
def editar_evento(request, pk):
    """Formulario para editar eventos existentes"""
    evento = get_object_or_404(Evento, pk=pk)
    
    # Detectar de dónde viene el usuario
    referrer = request.META.get('HTTP_REFERER', '')
    viene_de_lista = 'lista' in referrer or 'eventos' in referrer
    viene_de_dashboard = 'dashboard' in referrer or referrer.endswith('/')
    
    if request.method == 'POST':
        form = EventoForm(request.POST, instance=evento)
        if form.is_valid():
            form.save()
            messages.success(request, f'Evento "{evento.nombre}" actualizado exitosamente.')
            
            # Mejorar la redirección basada en el contexto
            next_url = request.POST.get('next') or request.GET.get('next')
            
            if next_url:
                # Si hay una URL específica de retorno
                return redirect(next_url)
            elif viene_de_lista:
                # Si viene de la lista de eventos, regresar allí
                return redirect('lista_eventos')
            elif viene_de_dashboard:
                # Si viene del dashboard, regresar allí
                return redirect('dashboard') 
            else:
                # Por defecto, ir al dashboard
                return redirect('dashboard')
        else:
            messages.error(request, 'Por favor corrija los errores en el formulario.')
    else:
        form = EventoForm(instance=evento)
    
    # Determinar URLs de retorno para el template
    context = {
        'form': form,
        'evento': evento,
        'viene_de_lista': viene_de_lista,
        'viene_de_dashboard': viene_de_dashboard,
        'next_url': request.GET.get('next', ''),
    }
    
    return render(request, 'eventos/editar_evento.html', context)

# Vista para detalle de evento
# Vista para detalle de evento (versión modal)
@login_required
def detalle_evento(request, pk):
    """Muestra el detalle completo de un evento para modal"""
    evento = get_object_or_404(Evento, pk=pk)
    evento.actualizar_estado_automatico()
    
    # Siempre usar el template del modal
    return render(request, 'eventos/detalle_evento_modal.html', {'evento': evento})

# Vista para lista de eventos
@login_required
def lista_eventos(request):
    """Lista todos los eventos con filtros y paginación AJAX"""
    form = FiltroEventosForm(request.GET or None)
    eventos = Evento.objects.select_related('municipio', 'creado_por').order_by('-fecha_evento')
    
    # Aplicar filtros
    if form.is_valid():
        # Filtro por fecha desde
        if form.cleaned_data.get('fecha_desde'):
            eventos = eventos.filter(fecha_evento__date__gte=form.cleaned_data['fecha_desde'])
        
        # Filtro por fecha hasta
        if form.cleaned_data.get('fecha_hasta'):
            eventos = eventos.filter(fecha_evento__date__lte=form.cleaned_data['fecha_hasta'])
        
        # Filtro por municipio
        if form.cleaned_data.get('municipio'):
            eventos = eventos.filter(municipio=form.cleaned_data['municipio'])
        
        # Filtro por estado
        if form.cleaned_data.get('estado'):
            eventos = eventos.filter(estado=form.cleaned_data['estado'])
        
        # Filtro por asistencia del gobernador
        asistencia = form.cleaned_data.get('asistencia')
        if asistencia:
            if asistencia == 'True':
                eventos = eventos.filter(asistio_gobernador=True)
            elif asistencia == 'False':
                eventos = eventos.filter(asistio_gobernador=False)
        
        # Filtro por tipo de evento
        tipo_evento = form.cleaned_data.get('tipo_evento')
        if tipo_evento:
            if tipo_evento == 'festivo':
                eventos = eventos.filter(es_festivo=True)
            elif tipo_evento == 'regular':
                eventos = eventos.filter(es_festivo=False)
        
        # Filtro por búsqueda de texto
        buscar = form.cleaned_data.get('buscar')
        if buscar:
            from django.db.models import Q
            eventos = eventos.filter(
                Q(nombre__icontains=buscar) |
                Q(lugar__icontains=buscar) |
                Q(responsable__icontains=buscar) |
                Q(municipio__nombre__icontains=buscar) |
                Q(representante__icontains=buscar) |
                Q(descripcion__icontains=buscar) |
                Q(observaciones__icontains=buscar)
            )
    
    # Implementar paginación
    paginator = Paginator(eventos, 4)  # 4 eventos por página
    page_number = request.GET.get('page', 1)
    
    try:
        page_obj = paginator.get_page(page_number)
    except PageNotAnInteger:
        page_obj = paginator.get_page(1)
    except EmptyPage:
        page_obj = paginator.get_page(paginator.num_pages)
    
    # Calcular estadísticas
    total_eventos = eventos.count()
    hoy = timezone.now().date()
    eventos_hoy = eventos.filter(fecha_evento__date=hoy)
    eventos_proximos = eventos.filter(fecha_evento__date__gt=hoy)
    eventos_finalizados = eventos.filter(estado='finalizado')
    
    # Info de paginación
    pagination_info = {
        'showing_start': (page_obj.number - 1) * 4 + 1 if page_obj.object_list else 0,
        'showing_end': min(page_obj.number * 4, total_eventos),
        'total_count': total_eventos,
        'current_page': page_obj.number,
        'total_pages': paginator.num_pages,
    }
    
    # NUEVO: Manejar peticiones AJAX
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        # Para peticiones AJAX, devolver solo el HTML de la tabla y paginación
        table_html = render_to_string('eventos/partials/events_table.html', {
            'eventos': page_obj.object_list,
            'page_obj': page_obj,
            'pagination_info': pagination_info,
            'request': request,
        })
        
        return JsonResponse({
            'success': True,
            'table_html': table_html,
            'pagination_info': pagination_info,
            'total_eventos': total_eventos,
        })
    
    # Para peticiones normales, renderizar la página completa
    context = {
        'eventos': page_obj.object_list,
        'page_obj': page_obj,
        'form': form,
        'total_eventos': total_eventos,
        'eventos_hoy': eventos_hoy,
        'eventos_proximos': eventos_proximos,
        'eventos_finalizados': eventos_finalizados,
        'pagination_info': pagination_info,
    }
    
    return render(request, 'eventos/lista_eventos.html', context)





# Vistas de debug
@login_required
def verificar_fechas_eventos(request):
    """Vista para debug de fechas de eventos"""
    mexico_tz = pytz.timezone('America/Mexico_City')
    ahora_mexico = timezone.now().astimezone(mexico_tz)
    hoy = ahora_mexico.date()
    
    eventos = Evento.objects.all().order_by('fecha_evento')
    
    info_eventos = []
    for evento in eventos:
        evento_fecha_mexico = evento.fecha_evento.astimezone(mexico_tz)
        info_eventos.append({
            'evento': evento,
            'fecha_utc': evento.fecha_evento,
            'fecha_mexico': evento_fecha_mexico,
            'fecha_solo': evento_fecha_mexico.date(),
            'es_hoy': evento_fecha_mexico.date() == hoy,
            'estado': evento.estado,
        })
    
    context = {
        'info_eventos': info_eventos,
        'ahora_mexico': ahora_mexico,
        'hoy': hoy,
    }
    
    return render(request, 'eventos/debug_fechas.html', context)

@login_required
def cambiar_fecha_evento_debug(request, pk):
    """Cambia la fecha de un evento para testing"""
    if request.method == 'POST':
        evento = get_object_or_404(Evento, pk=pk)
        nueva_fecha = request.POST.get('nueva_fecha')
        nueva_hora = request.POST.get('nueva_hora')
        
        if nueva_fecha and nueva_hora:
            try:
                mexico_tz = pytz.timezone('America/Mexico_City')
                fecha_naive = datetime.strptime(f"{nueva_fecha} {nueva_hora}", "%Y-%m-%d %H:%M")
                fecha_mexico = mexico_tz.localize(fecha_naive)
                
                evento.fecha_evento = fecha_mexico
                evento.save()
                
                messages.success(request, f'Fecha del evento "{evento.nombre}" cambiada a {fecha_mexico}')
            except ValueError as e:
                messages.error(request, f'Error en el formato de fecha/hora: {e}')
    
    return redirect('verificar_fechas_eventos')

@login_required
def crear_eventos_prueba(request):
    """Crea eventos de prueba para hoy"""
    if request.method == 'POST':
        mexico_tz = pytz.timezone('America/Mexico_City')
        ahora_mexico = timezone.now().astimezone(mexico_tz)
        hoy = ahora_mexico.date()
        
        # Crear evento en curso (hace 30 minutos)
        evento_en_curso = Evento.objects.create(
            nombre="Evento en Curso - Prueba",
            fecha_evento=mexico_tz.localize(datetime.combine(hoy, datetime.min.time().replace(hour=max(0, ahora_mexico.hour-1), minute=30))),
            municipio=Municipio.objects.first(),
            lugar="Lugar de prueba",
            responsable="Responsable de prueba",
            estado="en_curso",
            creado_por=request.user
        )
        
        # Crear evento próximo (en 2 horas)
        evento_proximo = Evento.objects.create(
            nombre="Evento Próximo - Prueba",
            fecha_evento=mexico_tz.localize(datetime.combine(hoy, datetime.min.time().replace(hour=min(23, ahora_mexico.hour+2), minute=0))),
            municipio=Municipio.objects.first(),
            lugar="Lugar de prueba 2",
            responsable="Responsable de prueba 2",
            estado="programado",
            creado_por=request.user
        )
        
        # Crear evento finalizado (hace 3 horas)
        evento_finalizado = Evento.objects.create(
            nombre="Evento Finalizado - Prueba",
            fecha_evento=mexico_tz.localize(datetime.combine(hoy, datetime.min.time().replace(hour=max(0, ahora_mexico.hour-3), minute=0))),
            municipio=Municipio.objects.first(),
            lugar="Lugar de prueba 3",
            responsable="Responsable de prueba 3",
            estado="finalizado",
            creado_por=request.user
        )
        
        messages.success(request, 'Se crearon 3 eventos de prueba para hoy.')
    
    return redirect('dashboard')

# Vista para reportes
@login_required
def reportes(request):
    """Genera reportes con filtros"""
    form = FiltroEventosForm(request.GET or None)
    eventos = Evento.objects.select_related('municipio', 'creado_por').order_by('-fecha_evento')
    
    if form.is_valid():
        if form.cleaned_data['fecha_desde']:
            eventos = eventos.filter(fecha_evento__date__gte=form.cleaned_data['fecha_desde'])
        
        if form.cleaned_data['fecha_hasta']:
            eventos = eventos.filter(fecha_evento__date__lte=form.cleaned_data['fecha_hasta'])
        
        if form.cleaned_data['municipio']:
            eventos = eventos.filter(municipio=form.cleaned_data['municipio'])
        
        if form.cleaned_data['estado']:
            eventos = eventos.filter(estado=form.cleaned_data['estado'])
        
        if form.cleaned_data['asistio_gobernador'] is not None:
            eventos = eventos.filter(asistio_gobernador=form.cleaned_data['asistio_gobernador'])
    
    # Estadísticas
    total_eventos = eventos.count()
    eventos_gobernador = eventos.filter(asistio_gobernador=True).count()
    eventos_representante = eventos.filter(asistio_gobernador=False).count()
    eventos_festivos = eventos.filter(es_festivo=True).count()
    
    return render(request, 'reportes/reportes.html', {
        'form': form,
        'eventos': eventos,
        'total_eventos': total_eventos,
        'eventos_gobernador': eventos_gobernador,
        'eventos_representante': eventos_representante,
        'eventos_festivos': eventos_festivos,
    })

# Vista para generar Excel
@login_required
def generar_excel(request):
    """Genera reporte en Excel"""
    form = FiltroEventosForm(request.GET or None)
    eventos = Evento.objects.select_related('municipio', 'creado_por').order_by('-fecha_evento')
    
    if form.is_valid():
        if form.cleaned_data['fecha_desde']:
            eventos = eventos.filter(fecha_evento__date__gte=form.cleaned_data['fecha_desde'])
        
        if form.cleaned_data['fecha_hasta']:
            eventos = eventos.filter(fecha_evento__date__lte=form.cleaned_data['fecha_hasta'])
        
        if form.cleaned_data['municipio']:
            eventos = eventos.filter(municipio=form.cleaned_data['municipio'])
        
        if form.cleaned_data['estado']:
            eventos = eventos.filter(estado=form.cleaned_data['estado'])
        
        if form.cleaned_data['asistio_gobernador'] is not None:
            eventos = eventos.filter(asistio_gobernador=form.cleaned_data['asistio_gobernador'])
    
    # Crear workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte de Eventos"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Headers
    headers = [
        'Evento', 'Fecha', 'Hora', 'Municipio', 'Lugar', 'Responsable',
        'Estado', 'Asistió Gobernador', 'Representante', 'Es Festivo', 'Creado Por'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Datos
    for row, evento in enumerate(eventos, 2):
        fecha_mexico = evento.get_fecha_mexico()
        ws.cell(row=row, column=1, value=evento.nombre)
        ws.cell(row=row, column=2, value=fecha_mexico.strftime('%d/%m/%Y'))
        ws.cell(row=row, column=3, value=fecha_mexico.strftime('%H:%M'))
        ws.cell(row=row, column=4, value=evento.municipio.nombre)
        ws.cell(row=row, column=5, value=evento.lugar)
        ws.cell(row=row, column=6, value=evento.responsable)
        ws.cell(row=row, column=7, value=evento.get_estado_display())
        ws.cell(row=row, column=8, value="Sí" if evento.asistio_gobernador else "No")
        ws.cell(row=row, column=9, value=evento.representante or "N/A")
        ws.cell(row=row, column=10, value="Sí" if evento.es_festivo else "No")
        ws.cell(row=row, column=11, value=evento.creado_por.get_full_name() or evento.creado_por.username)
    
    # Ajustar ancho de columnas
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Crear respuesta HTTP
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="reporte_eventos_{timezone.now().strftime("%Y%m%d_%H%M")}.xlsx"'
    
    # Guardar workbook en respuesta
    wb.save(response)
    return response

# Vista para estadísticas
@login_required
def estadisticas(request):
    """Muestra estadísticas generales"""
    # Estadísticas básicas
    total_eventos = Evento.objects.count()
    eventos_gobernador = Evento.objects.filter(asistio_gobernador=True).count()
    eventos_representante = Evento.objects.filter(asistio_gobernador=False).count()
    eventos_festivos = Evento.objects.filter(es_festivo=True).count()
    
    # Calcular porcentajes
    porcentaje_gobernador = round((eventos_gobernador / total_eventos * 100), 1) if total_eventos > 0 else 0
    porcentaje_representante = round((eventos_representante / total_eventos * 100), 1) if total_eventos > 0 else 0
    porcentaje_festivos = round((eventos_festivos / total_eventos * 100), 1) if total_eventos > 0 else 0
    
    # Eventos por estado
    eventos_por_estado = Evento.objects.values('estado').annotate(
        total=Count('id')
    ).order_by('-total')
    
    # Eventos por municipio (top 10 para gráfica de donut)
    eventos_por_municipio = Evento.objects.values('municipio__nombre').annotate(
        total=Count('id')
    ).order_by('-total')[:10]
    
    # Datos para gráfica de municipios (Doughnut)
    municipios_data = {
        'labels': [item['municipio__nombre'] for item in eventos_por_municipio],
        'data': [item['total'] for item in eventos_por_municipio]
    }
    
    # MEJORADO: Datos para gráfico de asistencia por municipio (Top 20 con mejor información)
    asistencia_por_municipio = Evento.objects.values('municipio__nombre').annotate(
        total_gobernador=Count('id', filter=Q(asistio_gobernador=True)),
        total_representante=Count('id', filter=Q(asistio_gobernador=False)),
        total_eventos=Count('id')
    ).filter(total_eventos__gt=0).order_by('-total_eventos')[:20]
    
    # Preparar datos mejorados para la gráfica de barras
    asistencia_data = {
        'labels': [item['municipio__nombre'] for item in asistencia_por_municipio],
        'gobernador': [item['total_gobernador'] for item in asistencia_por_municipio],
        'representante': [item['total_representante'] for item in asistencia_por_municipio],
        'totales': [item['total_eventos'] for item in asistencia_por_municipio]
    }
    
    context = {
        'total_eventos': total_eventos,
        'eventos_gobernador': eventos_gobernador,
        'eventos_representante': eventos_representante,
        'eventos_festivos': eventos_festivos,
        'porcentaje_gobernador': porcentaje_gobernador,
        'porcentaje_representante': porcentaje_representante,
        'porcentaje_festivos': porcentaje_festivos,
        'eventos_por_estado': eventos_por_estado,
        'eventos_por_municipio': eventos_por_municipio,
        'municipios_data': json.dumps(municipios_data),
        'asistencia_data': json.dumps(asistencia_data),
    }
    
    return render(request, 'estadisticas/estadisticas.html', context)


# Vistas para el calendario
@login_required
def calendario(request):
    """Vista para mostrar el calendario de eventos"""
    return render(request, 'eventos/calendario.html')

@login_required
def eventos_calendario_api(request):
    """API para obtener eventos del calendario en formato JSON"""
    # Obtener parámetros de fecha del request
    year = request.GET.get('year')
    month = request.GET.get('month')
    
    # Si no se especifica año/mes, usar el actual
    if not year or not month:
        from .utils import get_current_mexico_time
        now = get_current_mexico_time()
        year = now.year
        month = now.month
    else:
        year = int(year)
        month = int(month)
    
    # Crear fechas de inicio y fin del mes
    from datetime import datetime
    import calendar
    from .utils import get_mexico_timezone
    
    # Primer día del mes
    fecha_inicio = datetime(year, month, 1)
    
    # Último día del mes
    ultimo_dia = calendar.monthrange(year, month)[1]
    fecha_fin = datetime(year, month, ultimo_dia, 23, 59, 59)
    
    # Obtener eventos del mes con timezone awareness
    mexico_tz = get_mexico_timezone()
    fecha_inicio = mexico_tz.localize(fecha_inicio)
    fecha_fin = mexico_tz.localize(fecha_fin)
    
    eventos = Evento.objects.filter(
        fecha_evento__gte=fecha_inicio,
        fecha_evento__lte=fecha_fin
    ).select_related('municipio', 'creado_por').order_by('fecha_evento')
    
    # Preparar datos para el calendario
    eventos_data = []
    for evento in eventos:
        # Actualizar estado automáticamente
        evento.actualizar_estado_automatico()
        
        # Usar utilidad para formatear fecha
        from .utils import format_event_date
        fecha_formateada = format_event_date(evento)
        
        eventos_data.append({
            'id': evento.id,
            'title': evento.nombre,
            'date': fecha_formateada['date'],
            'time': fecha_formateada['time'],
            'location': f"{evento.lugar}, {evento.municipio.nombre}",
            'description': evento.descripcion or 'Sin descripción',
            'estado': evento.estado,
            'es_festivo': evento.es_festivo,
            'asistio_gobernador': evento.asistio_gobernador,
            'responsable': evento.responsable,
            'attendees': 0,
            'municipio': evento.municipio.nombre,
            'full_location': evento.lugar,
        })
    
    return JsonResponse({'eventos': eventos_data}) 

@login_required
@require_http_methods(["POST"])
def chatbot_api(request):
    """API endpoint para el chatbot"""
    try:
        # Obtener el mensaje del request
        data = json.loads(request.body)
        mensaje = data.get('mensaje', '').strip()
        
        if not mensaje:
            return JsonResponse({
                'success': False,
                'respuesta': 'Por favor, escribe una consulta.',
                'error': 'Mensaje vacío'
            })
        
        # Procesar la consulta con el chatbot
        chatbot = ChatbotAgenda()
        respuesta = chatbot.procesar_consulta(mensaje)
        
        # Log de la consulta (opcional, para mejorar el chatbot)
        print(f"CHATBOT - Usuario: {request.user.username} | Consulta: {mensaje}")
        
        return JsonResponse({
            'success': True,
            'respuesta': respuesta,
            'timestamp': timezone.now().isoformat()
        })
        
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'respuesta': 'Error en el formato de la consulta.',
            'error': 'JSON inválido'
        }, status=400)
        
    except Exception as e:
        print(f"ERROR CHATBOT: {str(e)}")
        return JsonResponse({
            'success': False,
            'respuesta': 'Ocurrió un error procesando tu consulta. Inténtalo de nuevo.',
            'error': str(e)
        }, status=500)

@login_required
def chatbot_test(request):
    """Vista de testing para el chatbot (solo en desarrollo)"""
    if not settings.DEBUG:
        return redirect('dashboard')
    
    return render(request, 'chatbot/test.html')
